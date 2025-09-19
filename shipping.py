# -*- coding: utf-8 -*-
"""
shipping.py — packaging, Excel export, email, and GitHub push.

What this version does:
- Builds XLSX with a clean last-column "Open" hyperlink per row (no local C: paths shown).
  * The visible path column is sanitized to show a relative path or filename only.
  * Hyperlinks prefer RAW_BASE_URL; else PUBLIC_BASE_URL; else fall back to file://
- Adds zip_today(): create a ZIP containing only today's screenshots per site.
- Keeps zip_last_7_days(), send_email(), git_commit_and_push().

Environment (Windows User variables recommended):
  GIT_REPO_DIR            e.g., C:/Users/you/Downloads/adscraper-full-code
  GIT_REMOTE_NAME         default 'origin'
  GIT_BRANCH              default 'main'

  SMTP_HOST               e.g., smtp.gmail.com
  SMTP_PORT               e.g., 587
  SMTP_USER               your_smtp_username
  SMTP_PASS               your_app_password
  MAIL_FROM               "Ad Bot <bot@example.com>"
  MAIL_TO                 comma-separated list (e.g., "a@b.com,c@d.com")

  PUBLIC_BASE_URL         e.g., https://github.com/<USER>/<REPO>/blob/main
  RAW_BASE_URL            e.g., https://raw.githubusercontent.com/<USER>/<REPO>/main  (optional but preferred)
  OUTPUT_ROOT             (optional) local screenshots root for computing relative paths
"""

import os, csv, zipfile, smtplib, mimetypes, traceback, subprocess
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ---------- helpers ----------

def _guess_output_root() -> str:
    """Prefer explicit env; else infer a sibling 'banner_screenshots' next to this file."""
    env_root = os.getenv("OUTPUT_ROOT", "").strip()
    if env_root:
        return os.path.abspath(env_root)
    here = Path(__file__).resolve().parent
    candidate = here / "banner_screenshots"
    return str(candidate)

def _to_rel(path_str: str, output_root: str) -> str:
    try:
        rel = os.path.relpath(path_str, output_root)
        return rel.replace("\\", "/")
    except Exception:
        return ""

def _file_url(local_path: str) -> str:
    # Excel will open file:// links when accessible to the viewer
    p = Path(local_path).resolve()
    return "file:///" + str(p).replace("\\", "/")


# ---------- Excel builder (clean links, no local C: paths) ----------

def build_xlsx_from_csv(csv_path: str, xlsx_path: str) -> None:
    """
    Build XLSX that hides local paths entirely:
      - EXCLUDES 'example_path' and 'image_path' columns from output
      - Appends an 'open' column with a clickable hyperlink per row (RAW > PUBLIC > file://)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "banners"

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

    if not os.path.exists(csv_path):
        wb.save(xlsx_path)
        print("[XLSX] CSV not found; created empty workbook.")
        return

    # Local screenshots root to compute relative paths
    output_root = _guess_output_root()

    # URL bases
    raw_base = (os.getenv("RAW_BASE_URL") or "").rstrip("/")
    public_base = (os.getenv("PUBLIC_BASE_URL") or "").rstrip("/")

    # If RAW not provided but PUBLIC is a GitHub blob URL, auto-derive RAW
    if not raw_base and public_base and "/blob/" in public_base:
        try:
            left, branch = public_base.split("/blob/", 1)
            raw_base = left.replace("https://github.com/", "https://raw.githubusercontent.com/") + "/" + branch
        except Exception:
            pass

    def best_public_url(rel_path: str) -> str:
        if not rel_path:
            return ""
        if raw_base:
            return raw_base.rstrip("/") + "/" + rel_path.lstrip("/")
        if public_base:
            return public_base.rstrip("/") + "/" + rel_path.lstrip("/")
        return ""

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        src_fields = reader.fieldnames or []

        # Exclude path columns from the Excel output
        path_cols = {"example_path", "image_path"}
        out_fields = [c for c in src_fields if c not in path_cols]
        # Add the 'open' column at the end
        out_fields.append("open")
        ws.append(out_fields)

        # Which source column to read path from (still read it from CSV; just not showing it)
        chosen_path_col = None
        for c in ("example_path", "image_path"):
            if c in src_fields:
                chosen_path_col = c
                break

        for row in reader:
            # Prepare visible row values (without path columns)
            visible_values = [row.get(c, "") for c in src_fields if c not in path_cols]
            # Placeholder for 'open' text
            visible_values.append("Open")
            ws.append(visible_values)

            # Build hyperlink
            original_path = (row.get(chosen_path_col, "") if chosen_path_col else "") or ""
            rel = _to_rel(original_path, output_root) if original_path else ""
            public_url = best_public_url(rel)
            file_link = public_url if public_url else (_file_url(original_path) if original_path else "")

            # Set hyperlink on the last cell (the 'open' column)
            open_cell = ws.cell(row=ws.max_row, column=len(out_fields))
            if file_link:
                open_cell.hyperlink = file_link
                open_cell.style = "Hyperlink"

        # Tidy column widths
        for i, col in enumerate(out_fields, 1):
            ws.column_dimensions[get_column_letter(i)].width = 28

    wb.save(xlsx_path)
    print(f"[XLSX] Wrote {xlsx_path} (only 'open' link; no local paths shown)")


# ---------- ZIP helpers ----------

def zip_today(root_screenshots: str, out_zip_path: str, sites=("gogo.mn", "ikon.mn", "news.mn"), day: str | None = None) -> bool:
    """
    Zip only today's (or given YYYY-MM-DD) screenshots per site into out_zip_path.
    Returns True if a zip was created with at least one file, else False.
    """
    day = day or datetime.now().strftime("%Y-%m-%d")
    os.makedirs(os.path.dirname(out_zip_path), exist_ok=True)
    count = 0
    with zipfile.ZipFile(out_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for site in sites:
            folder = os.path.join(root_screenshots, site, day)
            if not os.path.isdir(folder):
                continue
            for dirpath, _, filenames in os.walk(folder):
                for fn in filenames:
                    full = os.path.join(dirpath, fn)
                    rel = os.path.relpath(full, root_screenshots).replace("\\", "/")
                    zf.write(full, rel)
                    count += 1
    if count:
        print(f"[ZIP] Today-only zip wrote {out_zip_path} with {count} file(s)")
        return True
    print("[ZIP] No files for today; zip not created")
    return False


def zip_last_7_days(root_screenshots: str, out_zip_path: str) -> None:
    cutoff = datetime.now().date() - timedelta(days=7)
    os.makedirs(os.path.dirname(out_zip_path), exist_ok=True)
    with zipfile.ZipFile(out_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for site in ("gogo.mn", "ikon.mn", "news.mn"):
            site_root = os.path.join(root_screenshots, site)
            if not os.path.isdir(site_root):
                continue
            for day in os.listdir(site_root):
                # day folder like YYYY-MM-DD
                try:
                    d = datetime.strptime(day, "%Y-%m-%d").date()
                except Exception:
                    continue
                if d >= cutoff:
                    folder = os.path.join(site_root, day)
                    for dirpath, _, filenames in os.walk(folder):
                        for fn in filenames:
                            full = os.path.join(dirpath, fn)
                            rel  = os.path.relpath(full, root_screenshots)
                            zf.write(full, rel)
    print(f"[ZIP] Wrote {out_zip_path}")


# ---------- Email + Git ----------

def _attach_file(msg: EmailMessage, file_path: str):
    ctype, encoding = mimetypes.guess_type(file_path)
    if ctype is None or encoding is not None:
        ctype = "application/octet-stream"
    maintype, subtype = ctype.split("/", 1)
    with open(file_path, "rb") as f:
        msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=os.path.basename(file_path))

def send_email(subject: str, body: str, attachments: list[str]) -> None:
    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    mail_from = os.getenv("MAIL_FROM", smtp_user)
    mail_to   = [x.strip() for x in os.getenv("MAIL_TO", "").split(",") if x.strip()]

    if not (smtp_host and smtp_user and smtp_pass and mail_to):
        print("[WARN] Email not sent (SMTP env missing)")
        return

    msg = EmailMessage()
    msg["From"] = mail_from
    msg["To"] = ", ".join(mail_to)
    msg["Subject"] = subject
    msg.set_content(body)

    for p in attachments:
        try:
            if os.path.exists(p):
                _attach_file(msg, p)
        except Exception:
            traceback.print_exc()

    with smtplib.SMTP(smtp_host, smtp_port) as s:
        s.starttls()
        s.login(smtp_user, smtp_pass)
        s.send_message(msg)
    print("[MAIL] Sent email to:", msg["To"])

def git_commit_and_push(repo_dir: str, message: str) -> None:
    """
    Uses local git in PATH. Repo should already have remote + auth set.
    Env (optional): GIT_REMOTE_NAME, GIT_BRANCH
    """
    remote = os.getenv("GIT_REMOTE_NAME", "origin")
    branch = os.getenv("GIT_BRANCH", "main")
    def run(*cmd):
        subprocess.run(cmd, cwd=repo_dir, check=False)
    run("git", "add", "-A")
    run("git", "commit", "-m", message)
    run("git", "push", remote, branch)
    print("[GIT] Push attempted to", remote, branch)
